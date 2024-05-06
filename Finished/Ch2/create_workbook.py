# LinkedIn Learning Course
# Example file for Python: Working with Excel and Spreadsheet Data by Joe Marini
# Create a new workbook with worksheets and add content 

from openpyxl import Workbook
import datetime
import random

# Create a new workbook
wb = Workbook()

# Get the active worksheet and name it "TestSheet"
sheet = wb.active
sheet.title = "First"

# Add some data to the new sheet
sheet["A1"] = "Test Data"
sheet["B1"] = 123.4567
sheet["C1"] = datetime.datetime(2030, 4, 1)

# Use the cell() function to fill a row with values
for i in range(1, 11):
    sheet.cell(row = 5, column = i).value = random.randint(1, 50)

# Create a second worksheet
sheet2 = wb.create_sheet("Second")
sheet2.cell(row=2, column=2).value = "More Data"

# Use the append() function to add rows to the end of the sheet
sheet2.append(["One","Two","Three"])
sheet2.append(["One","Two","Three"])
sheet2.append(["One","Two","Three"])

# Save the workbook - values don't update until we do this!
wb.save("NewWorkbook.xlsx")

print("Workbook created successfully!")
