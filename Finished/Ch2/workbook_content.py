# LinkedIn Learning Course
# Example file for Python: Working with Excel and Spreadsheet Data by Joe Marini
# Manipulate workbook content 

import openpyxl
from openpyxl.comments import Comment
from collections import defaultdict


# Create a new workbook
filename = "FinancialSample.xlsx"

# Load the workbook
wb = openpyxl.load_workbook(filename)

# Get the active worksheet
sheet = wb.active

# Get entire column or row of cells
col = sheet["C"]
row = sheet[10]
print(f"{len(col)} cells in column")
print(f"{len(row)} cells in row")

# Get a range of cells
range = sheet["A2:B7"]
print(f"{len(range)} items in range")
print(range)

# iterate over rows and columns
for col in sheet.iter_cols(min_row=2,max_row=3,min_col=2,max_col=5):
    for cell in col:
        print(cell.value)

counter = defaultdict(int)
for row in sheet.iter_rows(min_row=2,min_col=3,max_col=3):
    for cell in row:
        counter[cell.value] += 1
print(counter)


# create a cell with a comment in it
cell = sheet["A1"]
cell.comment = Comment("This is a comment", "Comment Author")

# save the workbook
wb.save("Content.xlsx")
