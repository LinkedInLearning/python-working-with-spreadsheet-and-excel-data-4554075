# LinkedIn Learning Course
# Example file for Python: Working with Excel and Spreadsheet Data by Joe Marini
# Open, load, and explore workbook content 

import openpyxl

filename = "FinancialSample.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(filename)

# Print basic information
print(f"Number of worksheets: {len(workbook.sheetnames)}")

# Explore each worksheet
for worksheet_name in workbook.sheetnames:
    worksheet = workbook[worksheet_name]
    print(f"\nWorksheet: {worksheet_name}")
    
    # Get dimensions
    dimensions = worksheet.dimensions
    print(f"  - Dimensions: {dimensions}")

    print(f"Min row: {worksheet.min_row}")
    print(f"Max row: {worksheet.max_row}")
    print(f"Min column: {worksheet.min_column}")
    print(f"Max column: {worksheet.max_column}")

    # Check if the worksheet is empty
    if worksheet.max_row == 1 and worksheet.max_column == 1:
        print("  - Worksheet is empty")
    else:
        # Get a cell (e.g., top-left corner)
        cell = worksheet["A1"]
        print(f"  - Top-left cell value: {cell.value}")

        cell = worksheet.cell(row=worksheet.max_row,column=worksheet.max_column)
        print(f"  - Bottom-Right cell value: {cell.value}")
