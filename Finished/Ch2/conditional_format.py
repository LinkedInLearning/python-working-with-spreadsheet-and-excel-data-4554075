# LinkedIn Learning Course
# Example file for Python: Working with Excel and Spreadsheet Data by Joe Marini
# Apply conditional formatting to a worksheet 

import openpyxl
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle


filename = "FinancialSample.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(filename)
sheet = workbook["SalesData"]

# define the style to represent the formatting
red_color = "ffd2d2"
bold_text = Font(bold=True, color="00FF0000")
red_fill = PatternFill(bgColor=red_color, fill_type='solid')
diff_style = DifferentialStyle(font=bold_text, fill=red_fill)

# create a rule for the condition
rule = Rule(type="expression", dxf=diff_style, formula=["$L1<10000"])

# add the rule to the entire sheet
dimensions = sheet.dimensions
sheet.conditional_formatting.add(dimensions, rule)

workbook.save("CondFormat.xlsx")
print("Workbook created successfully!")
